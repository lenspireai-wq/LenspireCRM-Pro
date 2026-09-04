from django.db import transaction
from django.utils import timezone
from datetime import datetime, time
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from rest_framework import serializers, status, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from apps.core.api import OrganizationScopedViewSet
from apps.core.permissions import SalesAccessPermission, SharedBookingAccessPermission
from apps.operations.models import CalendarEvent
from apps.production.models import ProductionJob
from .models import Booking, Customer, Lead, LeadActivity, SalesTarget

LEAD_ADVANCE_NOTE = "Advance Booking amount recorded during lead confirmation."


class LeadSerializer(serializers.ModelSerializer):
    activities = serializers.SerializerMethodField()
    attachments = serializers.SerializerMethodField()
    class Meta: model = Lead; fields = "__all__"; read_only_fields = ("organization", "lead_code")
    def get_activities(self, obj):
        return LeadActivitySerializer(obj.activities.order_by("-created_at", "-id"), many=True).data
    def get_attachments(self, obj):
        request = self.context.get("request")
        return [{"id": item.id, "name": item.name, "file": request.build_absolute_uri(item.file.url) if request else item.file.url, "created_at": item.created_at} for item in obj.attachments.order_by("-created_at")]
    def validate(self, attrs):
        status_value = attrs.get("status", getattr(self.instance, "status", "New"))
        lost_reason = attrs.get("lost_reason", getattr(self.instance, "lost_reason", ""))
        if status_value == "Lost" and not str(lost_reason).strip():
            raise serializers.ValidationError({"lost_reason": "Select a reason when marking a lead as Lost."})
        if status_value != "Lost":
            attrs["lost_reason"] = ""
        event_date = attrs.get("event_date", getattr(self.instance, "event_date", None))
        if not event_date:
            raise serializers.ValidationError({"event_date": "Event date is required."})
        next_followup = attrs.get("next_followup_at", getattr(self.instance, "next_followup_at", None))
        if status_value == "Follow-up" and not next_followup:
            raise serializers.ValidationError({"next_followup_at": "Schedule the next follow-up date and time."})
        organization = self.context["request"].user.organization
        mobile = str(attrs.get("mobile", getattr(self.instance, "mobile", ""))).strip()
        normalized = "".join(character for character in mobile if character.isdigit())[-10:]
        if normalized:
            candidates = Lead.objects.filter(organization=organization).exclude(pk=getattr(self.instance, "pk", None))
            duplicate = next((lead for lead in candidates.only("id", "name", "mobile", "lead_code") if "".join(c for c in lead.mobile if c.isdigit())[-10:] == normalized), None)
            if duplicate:
                raise serializers.ValidationError({"mobile": f"Duplicate mobile number already belongs to {duplicate.name} ({duplicate.lead_code})."})
        else:
            name = str(attrs.get("name", getattr(self.instance, "name", ""))).strip()
            event_type = str(attrs.get("event_type", getattr(self.instance, "event_type", ""))).strip()
            event_date = attrs.get("event_date", getattr(self.instance, "event_date", None))
            duplicate = Lead.objects.filter(
                organization=organization,
                name__iexact=name,
                event_type__iexact=event_type,
                event_date=event_date,
            ).exclude(pk=getattr(self.instance, "pk", None)).first()
            if duplicate:
                raise serializers.ValidationError({"detail": f"This lead already exists as {duplicate.name} ({duplicate.lead_code})."})
        if status_value in {"Booked", "Confirmed"}:
            required = ("couple_name", "total_closing", "payment_mode", "advance_received", "received_by", "payment_received_date")
            missing = [field for field in required if attrs.get(field, getattr(self.instance, field, None)) in (None, "")]
            if missing: raise serializers.ValidationError({field: "Required for a booked lead." for field in missing})
            total_closing = attrs.get("total_closing", getattr(self.instance, "total_closing", 0)) or 0
            advance_received = attrs.get("advance_received", getattr(self.instance, "advance_received", 0)) or 0
            if total_closing <= 0:
                raise serializers.ValidationError({"total_closing": "Total closing must be greater than zero."})
            if advance_received <= 0:
                raise serializers.ValidationError({"advance_received": "Advance received must be greater than zero."})
            if advance_received > total_closing:
                raise serializers.ValidationError({"advance_received": "Advance cannot exceed the total closing amount."})
        return attrs

class LeadActivitySerializer(serializers.ModelSerializer):
    class Meta: model = LeadActivity; fields = "__all__"; read_only_fields = ("organization", "lead")
class CustomerSerializer(serializers.ModelSerializer):
    class Meta: model = Customer; fields = "__all__"; read_only_fields = ("organization",)
class BookingSerializer(serializers.ModelSerializer):
    class Meta: model = Booking; fields = "__all__"; read_only_fields = ("organization",)
class SalesTargetSerializer(serializers.ModelSerializer):
    class Meta: model = SalesTarget; fields = "__all__"; read_only_fields = ("organization",)

class LeadViewSet(OrganizationScopedViewSet):
    queryset = Lead.objects.all().order_by("-created_at")
    serializer_class = LeadSerializer
    permission_classes = (SalesAccessPermission,)
    filterset_fields = {"status": ["exact", "in"], "priority": ["exact", "in"], "event_type": ["exact", "in"], "source": ["exact", "in"], "city": ["exact", "icontains"], "assigned_to": ["exact", "icontains"], "event_date": ["exact", "gte", "lte"], "next_followup_at": ["gte", "lte"], "created_at": ["gte", "lte"]}
    search_fields = ("lead_code", "name", "mobile", "client_name", "client_mobile", "couple_name", "city", "notes", "referred_by", "referral_code")
    ordering_fields = ("created_at", "event_date", "next_followup_at", "total_closing", "priority")
    throttle_scope = "lead_write"
    @transaction.atomic
    def perform_create(self, serializer):
        organization = self.request.user.organization
        from apps.core.models import Organization
        Organization.objects.select_for_update().get(pk=organization.pk)
        numbers = [int(code[1:]) for code in Lead.objects.filter(organization=organization, lead_code__regex=r"^L[0-9]+$").values_list("lead_code", flat=True)]
        lead = serializer.save(organization=organization, lead_code=f"L{max(numbers, default=0) + 1:03d}")
        LeadActivity.objects.create(organization=organization, lead=lead, activity_type="Lead Created", description=f"Lead created with status {lead.status}.", performed_by=self.request.user.display_name or self.request.user.username)
        try:
            from apps.notifications.models import broadcast
            broadcast(
                organization=organization,
                title=f"New lead: {lead.name}",
                body=f"{lead.event_type} on {lead.event_date or 'TBD'} · {lead.city or '—'}",
                level="info",
                category="sales",
                link="/sales",
                payload={"lead_id": lead.id},
            )
        except Exception:
            pass
        if lead.status in {"Booked", "Confirmed"}: self._convert(lead)
    @transaction.atomic
    def perform_update(self, serializer):
        previous = self.get_object(); old_status = previous.status; old_followup = previous.next_followup_at
        lead = serializer.save()
        actor = self.request.user.display_name or self.request.user.username
        if old_status != lead.status:
            LeadActivity.objects.create(organization=lead.organization, lead=lead, activity_type="Status Change", description=f"Status changed from {old_status} to {lead.status}.", performed_by=actor)
        if old_followup != lead.next_followup_at:
            description = f"Next follow-up scheduled for {lead.next_followup_at}." if lead.next_followup_at else "Follow-up schedule cleared."
            LeadActivity.objects.create(organization=lead.organization, lead=lead, activity_type="Follow-up", description=description, performed_by=actor)
        if lead.status in {"Booked", "Confirmed"}: self._convert(lead)
        self._sync_connected(lead)
    def perform_destroy(self, instance):
        if Customer.objects.filter(lead=instance).exists() or Booking.objects.filter(lead=instance).exists():
            raise serializers.ValidationError("Converted leads cannot be deleted because they are linked to customer and booking records.")
        instance.delete()
    def _convert(self, lead):
        from apps.accounts.models import Payment

        org = lead.organization
        customer, _ = Customer.objects.get_or_create(organization=org, lead=lead, defaults={"customer_code":f"CUS-{lead.id:05d}","name":lead.name,"phone":lead.mobile,"city":lead.city,"source":lead.source})
        booking, _ = Booking.objects.get_or_create(organization=org, lead=lead, defaults={"booking_code":f"BKG-{lead.id:05d}","customer":customer,"event_type":lead.event_type,"event_date":lead.event_date,"city":lead.city,"quoted_amount":lead.total_closing or lead.budget or 0})
        if lead.advance_received and lead.advance_received > 0:
            paid_at = timezone.make_aware(datetime.combine(lead.payment_received_date, time.min)) if lead.payment_received_date else timezone.now()
            Payment.objects.update_or_create(
                organization=org,
                booking=booking,
                notes=LEAD_ADVANCE_NOTE,
                defaults={
                    "customer": customer,
                    "payment_type": "Advance",
                    "amount": lead.advance_received,
                    "status": "Paid",
                    "payment_mode": lead.payment_mode,
                    "received_by": lead.received_by,
                    "paid_at": paid_at,
                },
            )
        ProductionJob.objects.get_or_create(organization=org, booking=booking, defaults={"customer":customer,"due_date":lead.event_date})
        if lead.event_date:
            CalendarEvent.objects.update_or_create(
                organization=org,
                booking=booking,
                defaults={
                    "customer": customer,
                    "title": f"{lead.name} · {lead.event_type}",
                    "client_name": lead.client_name or lead.name,
                    "handled_by": lead.assigned_to,
                    "couple_name": lead.couple_name,
                    "contact_no": lead.client_mobile or lead.mobile,
                    "event_type": lead.event_type,
                    "start_date": lead.event_date,
                    "city": lead.city,
                    "notes": lead.notes,
                },
            )
        if lead.status != "Confirmed": lead.status="Confirmed"; lead.save(update_fields=("status","updated_at"))
        return customer, booking
    def _sync_connected(self, lead):
        customer = Customer.objects.filter(lead=lead).first()
        if not customer: return
        customer.name=lead.name; customer.phone=lead.mobile; customer.city=lead.city; customer.source=lead.source; customer.save()
        booking = Booking.objects.filter(lead=lead).first()
        if booking:
            booking.event_type=lead.event_type; booking.event_date=lead.event_date; booking.city=lead.city; booking.quoted_amount=lead.total_closing or lead.budget or 0; booking.save()
            CalendarEvent.objects.filter(booking=booking).update(
                title=f"{lead.name} · {lead.event_type}",
                client_name=lead.client_name or lead.name,
                handled_by=lead.assigned_to,
                couple_name=lead.couple_name,
                contact_no=lead.client_mobile or lead.mobile,
                event_type=lead.event_type,
                start_date=lead.event_date,
                city=lead.city,
                notes=lead.notes,
            )
            ProductionJob.objects.filter(booking=booking).update(due_date=lead.event_date)
    @action(detail=True, methods=["post"], url_path="activities")
    def add_activity(self, request, pk=None):
        lead = self.get_object(); activity_type = request.data.get("activity_type") or request.data.get("type")
        if activity_type not in {"Call","WhatsApp","Meeting","Note"}: return Response({"detail":"Select a valid activity type."}, status=status.HTTP_400_BAD_REQUEST)
        description = str(request.data.get("description", "")).strip()
        if not description: return Response({"detail":"Activity details are required."}, status=status.HTTP_400_BAD_REQUEST)
        activity = LeadActivity.objects.create(organization=lead.organization,lead=lead,activity_type=activity_type,description=description,performed_by=request.user.display_name or request.user.username)
        return Response(LeadActivitySerializer(activity).data, status=status.HTTP_201_CREATED)
    @action(detail=False, methods=["get"], url_path="export")
    def export_excel(self, request):
        workbook=Workbook(); sheet=workbook.active; sheet.title="Leads"
        sheet.append(["Lead Code","Customer Name","Mobile","Event Type","Event Date","City","Source","Status","Budget","Priority","Assigned To","Next Follow-up","Lost Reason","Referred By","Referral Code","Notes"])
        for lead in self.get_queryset().order_by("-created_at"):
            sheet.append([lead.lead_code,lead.name,lead.mobile,lead.event_type,lead.event_date.isoformat() if lead.event_date else "",lead.city,lead.source,lead.status,lead.budget,lead.priority,lead.assigned_to,lead.next_followup_at.isoformat() if lead.next_followup_at else "",lead.lost_reason,lead.referred_by,lead.referral_code,lead.notes])
        response=HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"]='attachment; filename="Lenspire-Leads.xlsx"'; workbook.save(response); return response
    @action(detail=False, methods=["get"], url_path="import-template")
    def import_template(self, request):
        format_ = (request.query_params.get("as") or "csv").lower()
        headers = [
            "name", "mobile", "event_type", "event_date", "city", "source",
            "status", "priority", "budget", "assigned_to", "next_followup_at",
            "referred_by", "referral_code", "notes",
        ]
        if format_ == "xlsx":
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Leads"
            sheet.append(headers)
            sheet.append([
                "Riya Sharma", "9876543210", "Wedding", "2026-12-15",
                "Mumbai", "Instagram", "New", "High", "250000",
                "Ananya", "2026-10-01T10:00:00", "Karan", "RIYA10",
                "Sample row — replace with your data",
            ])
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="Lenspire-Leads-Template.xlsx"'
            workbook.save(response)
            return response
        import csv
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="Lenspire-Leads-Template.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerow([
            "Riya Sharma", "9876543210", "Wedding", "2026-12-15",
            "Mumbai", "Instagram", "New", "High", "250000",
            "Ananya", "2026-10-01T10:00:00", "Karan", "RIYA10",
            "Sample row — replace with your data",
        ])
        return response

    @action(detail=False, methods=["post"], url_path="import")
    @transaction.atomic
    def import_excel(self, request):
        upload = request.FILES.get("file")
        if not upload:
            return Response({"detail": "Select an Excel or CSV file."}, status=status.HTTP_400_BAD_REQUEST)
        dry_run = str(request.data.get("dry_run", "")).lower() in {"1", "true", "yes"}
        name = (upload.name or "").lower()
        try:
            if name.endswith(".csv"):
                import csv
                import io
                text = io.TextIOWrapper(upload.file, encoding="utf-8-sig", newline="")
                reader = csv.reader(text)
                rows = list(reader)
                if not rows:
                    return Response({"detail": "The file is empty."}, status=status.HTTP_400_BAD_REQUEST)
                headers = [str(value or "").strip() for value in rows[0]]
                records = rows[1:]
            else:
                sheet = load_workbook(upload, read_only=True, data_only=True).active
                rows = list(sheet.iter_rows(values_only=True))
                if not rows:
                    return Response({"detail": "The file is empty."}, status=status.HTTP_400_BAD_REQUEST)
                headers = [str(value or "").strip() for value in rows[0]]
                records = [list(row) for row in rows[1:]]
        except Exception:
            return Response(
                {"detail": "The uploaded file is not a valid Excel workbook or CSV."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        normalize = lambda value: "".join(c for c in str(value or "").lower() if c.isalnum())
        normalized_headers = [normalize(value) for value in headers]
        aliases = {
            "customername": "name", "name": "name", "mobile": "mobile", "mobilenumber": "mobile",
            "eventtype": "event_type", "event": "event_type", "eventdate": "event_date",
            "city": "city", "source": "source", "status": "status", "budget": "budget",
            "priority": "priority", "assignedto": "assigned_to", "nextfollowup": "next_followup_at",
            "lostreason": "lost_reason", "referredby": "referred_by", "referralcode": "referral_code",
            "notes": "notes", "couplename": "couple_name", "totalclosing": "total_closing",
            "paymentmode": "payment_mode", "advancereceived": "advance_received",
            "receivedby": "received_by", "paymentreceiveddate": "payment_received_date",
        }
        mapping = []
        unmapped = []
        for original, key in zip(headers, normalized_headers):
            if key in aliases:
                mapping.append({"header": original, "field": aliases[key]})
            else:
                unmapped.append(original)
        report = {
            "imported": 0,
            "skipped": 0,
            "skipped_duplicates": 0,
            "skipped_invalid": 0,
            "dry_run": dry_run,
            "headers_detected": mapping,
            "headers_unmapped": unmapped,
            "errors": [],
        }
        organization = request.user.organization
        if dry_run:
            existing_phones = {
                "".join(c for c in lead.mobile if c.isdigit())[-10:]
                for lead in self.get_queryset().only("mobile")
            }
        else:
            existing_phones = set()
        for index, values in enumerate(records, start=2):
            data = {}
            for col_index, value in enumerate(values):
                if col_index >= len(normalized_headers):
                    break
                header_key = normalized_headers[col_index]
                if header_key in aliases and value not in (None, ""):
                    data[aliases[header_key]] = value
            row_label = (str(data.get("name", "")).strip() or f"Row {index}")
            if hasattr(data.get("event_date"), "date"):
                data["event_date"] = data["event_date"].date().isoformat()
            if hasattr(data.get("next_followup_at"), "isoformat"):
                data["next_followup_at"] = data["next_followup_at"].isoformat()
            if hasattr(data.get("payment_received_date"), "date"):
                data["payment_received_date"] = data["payment_received_date"].date().isoformat()
            if not str(data.get("name", "")).strip():
                report["skipped"] += 1
                report["skipped_invalid"] += 1
                report["errors"].append({"row": index, "name": row_label, "error": "Name is required."})
                continue
            mobile = str(data.get("mobile", "")).strip()
            normalized_mobile = "".join(c for c in mobile if c.isdigit())[-10:]
            if normalized_mobile and normalized_mobile in existing_phones:
                report["skipped"] += 1
                report["skipped_duplicates"] += 1
                report["errors"].append({"row": index, "name": row_label, "error": f"Duplicate of an existing lead ({mobile})."})
                continue
            data.update({
                "event_type": data.get("event_type") or "Other",
                "source": data.get("source") or "Excel Import",
                "status": data.get("status") if data.get("status") in {"New", "Follow-up", "Confirmed", "Lost"} else "New",
                "priority": data.get("priority") if data.get("priority") in {"High", "Medium", "Low"} else "Medium",
            })
            serializer = self.get_serializer(data=data)
            if not serializer.is_valid():
                report["skipped"] += 1
                report["skipped_invalid"] += 1
                first_error = next(iter(serializer.errors.values()), ["Invalid row"])[0]
                report["errors"].append({"row": index, "name": row_label, "error": str(first_error)})
                continue
            if dry_run:
                report["imported"] += 1
                if normalized_mobile:
                    existing_phones.add(normalized_mobile)
                continue
            try:
                self.perform_create(serializer)
                report["imported"] += 1
                if normalized_mobile:
                    existing_phones.add(normalized_mobile)
            except Exception as exc:
                report["skipped"] += 1
                report["skipped_invalid"] += 1
                report["errors"].append({"row": index, "name": row_label, "error": str(exc)})
        report["errors"] = report["errors"][:50]
        return Response(report)
    @action(detail=True, methods=["post"])
    @transaction.atomic
    def convert(self, request, pk=None):
        lead = self.get_object(); customer, booking = self._convert(lead)
        lead.status = "Confirmed"; lead.save(update_fields=("status", "updated_at"))
        LeadActivity.objects.create(organization=lead.organization, lead=lead, activity_type="Converted", description="Lead converted to connected workflow.", performed_by=request.user.display_name or request.user.username)
        return Response({"customer": CustomerSerializer(customer).data, "booking": BookingSerializer(booking).data})

class CustomerViewSet(OrganizationScopedViewSet):
    queryset = Customer.objects.all().order_by("-created_at")
    serializer_class = CustomerSerializer
    permission_classes = (SharedBookingAccessPermission,)
    filterset_fields = {"city": ["exact", "icontains"], "source": ["exact", "in"], "created_at": ["gte", "lte"]}
    search_fields = ("customer_code", "name", "phone", "email", "city")
    ordering_fields = ("created_at", "name")

class BookingViewSet(OrganizationScopedViewSet):
    queryset = Booking.objects.all().order_by("-event_date")
    serializer_class = BookingSerializer
    permission_classes = (SharedBookingAccessPermission,)
    filterset_fields = {"status": ["exact", "in"], "event_type": ["exact", "in"], "city": ["exact", "icontains"], "event_date": ["exact", "gte", "lte"], "quoted_amount": ["gte", "lte"], "created_at": ["gte", "lte"]}
    search_fields = ("booking_code", "package_name", "customer__name", "customer__phone", "lead__name", "lead__mobile", "city")
    ordering_fields = ("event_date", "quoted_amount", "created_at", "status")

class SalesTargetViewSet(OrganizationScopedViewSet):
    queryset = SalesTarget.objects.all().order_by("-target_month","salesperson")
    serializer_class = SalesTargetSerializer
    permission_classes = (SalesAccessPermission,)
    filterset_fields = {"target_month": ["exact", "gte", "lte"], "salesperson": ["exact", "icontains"]}
    search_fields = ("salesperson",)
    ordering_fields = ("target_month", "target_amount", "target_bookings")
    def perform_create(self, serializer):
        target, _ = SalesTarget.objects.update_or_create(organization=self.request.user.organization, salesperson=serializer.validated_data["salesperson"], target_month=serializer.validated_data["target_month"], defaults={"target_amount":serializer.validated_data.get("target_amount",0),"target_bookings":serializer.validated_data.get("target_bookings",0)})
        serializer.instance = target
