from datetime import date
from io import BytesIO

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from openpyxl import Workbook, load_workbook
from rest_framework.test import APIClient
from apps.core.models import Organization
from apps.sales.models import Customer, Booking
from apps.users.models import User
from .models import Payment, PaymentReminder


class AccountsTests(TestCase):
    def setUp(self):
        self.org = Organization.objects.create(name="Studio", slug="studio")
        self.user = User.objects.create_user(username="accounts", organization=self.org)
        self.customer = Customer.objects.create(organization=self.org, customer_code="C1", name="Test Client")
        self.booking = Booking.objects.create(organization=self.org, customer=self.customer, booking_code="B1", event_type="Wedding", quoted_amount=10000)
        self.client = APIClient()
        self.client.force_authenticate(self.user)

    def payload(self):
        return {"booking": self.booking.id, "customer": self.customer.id, "amount": "1000.00", "status": "Paid", "payment_type": "Advance"}

    def test_record_edit_delete_payment(self):
        response = self.client.post("/api/payments/", self.payload())
        self.assertEqual(response.status_code, 201, response.data)
        payment = Payment.objects.get()
        self.assertEqual(payment.organization, self.org)
        self.assertIsNotNone(payment.paid_at)
        self.assertEqual(response.data["client_name"], "Test Client")
        self.assertEqual(self.client.patch(f"/api/payments/{payment.id}/", {"amount": "1500.00"}).status_code, 200)
        self.assertEqual(self.client.delete(f"/api/payments/{payment.id}/").status_code, 204)

    def test_reject_invalid_amount_and_customer(self):
        payload = self.payload()
        payload["amount"] = "-1"
        self.assertEqual(self.client.post("/api/payments/", payload).status_code, 400)
        other = Organization.objects.create(name="Other", slug="other")
        customer = Customer.objects.create(organization=other, customer_code="C2", name="Other Client")
        payload = self.payload()
        payload["customer"] = customer.id
        self.assertEqual(self.client.post("/api/payments/", payload).status_code, 400)

    def test_reject_duplicate_payment(self):
        self.assertEqual(self.client.post("/api/payments/", self.payload()).status_code, 201)
        response = self.client.post("/api/payments/", self.payload())
        self.assertEqual(response.status_code, 400)
        self.assertIn("detail", response.data)

    def test_rejects_overpayment_and_excess_refund(self):
        overpayment = self.payload()
        overpayment["amount"] = "10001.00"
        response = self.client.post("/api/payments/", overpayment)
        self.assertEqual(response.status_code, 400)
        self.assertIn("amount", response.data)
        self.assertEqual(self.client.post("/api/payments/", self.payload()).status_code, 201)
        refund = self.payload()
        refund.update(amount="1001.00", payment_type="Refund")
        response = self.client.post("/api/payments/", refund)
        self.assertEqual(response.status_code, 400)
        self.assertIn("amount", response.data)

    def test_only_one_pending_entry_per_milestone(self):
        pending = self.payload()
        pending.update(status="Pending", due_date="2026-09-20")
        self.assertEqual(self.client.post("/api/payments/", pending).status_code, 201)
        pending["due_date"] = "2026-09-21"
        response = self.client.post("/api/payments/", pending)
        self.assertEqual(response.status_code, 400)
        self.assertIn("payment_type", response.data)

    def test_marking_pending_payment_paid_sets_payment_date(self):
        pending = self.payload()
        pending.update(status="Pending", due_date="2026-09-20")
        created = self.client.post("/api/payments/", pending)
        response = self.client.patch(
            f"/api/payments/{created.data['id']}/", {"status": "Paid"}
        )
        self.assertEqual(response.status_code, 200, response.data)
        self.assertIsNotNone(response.data["paid_at"])

    def test_daily_reminder_is_logged_once_and_updated(self):
        payload = {
            "booking": self.booking.id,
            "customer": self.customer.id,
            "payment_type": "Wedding Day",
            "milestone_amount": "4000.00",
            "outstanding_amount": "9000.00",
            "action": "Copied",
            "next_followup_date": "2026-09-06",
        }
        self.assertEqual(self.client.post("/api/payment-reminders/", payload).status_code, 201)
        payload.update(action="WhatsApp", next_followup_date="2026-09-07")
        response = self.client.post("/api/payment-reminders/", payload)
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(PaymentReminder.objects.count(), 1)
        reminder = PaymentReminder.objects.get()
        self.assertEqual(reminder.action, "WhatsApp")
        self.assertEqual(str(reminder.next_followup_date), "2026-09-07")

    def test_export_import(self):
        self.client.post("/api/payments/", self.payload())
        response = self.client.get("/api/payments/export/")
        self.assertEqual(response.status_code, 200)
        sheet = load_workbook(BytesIO(response.content)).active
        self.assertEqual(sheet.cell(2,1).value, "B1")
        workbook = Workbook()
        workbook.active.append(["booking_code", "amount", "status", "due_date"])
        workbook.active.append(["B1", 2000, "Pending", date(2026,10,1)])
        stream = BytesIO()
        workbook.save(stream)
        response = self.client.post("/api/payments/import/", {"file": SimpleUploadedFile("payments.xlsx",stream.getvalue())}, format="multipart")
        self.assertEqual(response.status_code, 201, response.data)
        self.assertEqual(Payment.objects.count(), 2)
