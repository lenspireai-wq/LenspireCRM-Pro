from io import BytesIO

from django.http import HttpResponse
from django.db import transaction
from django.db.models import Q, Sum
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from rest_framework import serializers
from rest_framework.decorators import action
from rest_framework.response import Response
from apps.core.api import OrganizationScopedViewSet
from apps.core.permissions import AccountsAccessPermission
from apps.sales.models import Booking, Customer
from .models import Payment, PaymentReminder

class PaymentSerializer(serializers.ModelSerializer):
    client_name = serializers.CharField(source="customer.name", read_only=True)
    booking_code = serializers.CharField(source="booking.booking_code", read_only=True)

    class Meta:
        model = Payment
        fields = "__all__"
        read_only_fields = ("organization",)

    def validate(self, attrs):
        organization = self.context["request"].user.organization
        booking = attrs.get("booking", getattr(self.instance, "booking", None))
        customer = attrs.get("customer", getattr(self.instance, "customer", None))
        if not organization:
            raise serializers.ValidationError("An organization is required to record payments.")
        if booking and booking.organization_id != organization.id:
            raise serializers.ValidationError({"booking": "Invalid booking."})
        if customer and customer.organization_id != organization.id:
            raise serializers.ValidationError({"customer": "Invalid customer."})
        if booking and customer and booking.customer_id != customer.id:
            raise serializers.ValidationError({"customer": "Customer does not match the booking."})
        if attrs.get("amount", getattr(self.instance, "amount", 0)) <= 0:
            raise serializers.ValidationError({"amount": "Enter an amount greater than zero."})
        amount = attrs.get("amount", getattr(self.instance, "amount", 0))
        payment_type = attrs.get("payment_type", getattr(self.instance, "payment_type", "Advance"))
        payment_status = attrs.get("status", getattr(self.instance, "status", "Pending"))
        if booking and payment_status == "Paid":
            existing = Payment.objects.filter(organization=organization, booking=booking, status="Paid")
            if self.instance:
                existing = existing.exclude(pk=self.instance.pk)
            totals = existing.aggregate(
                collected=Sum("amount", filter=~Q(payment_type="Refund")),
                refunded=Sum("amount", filter=Q(payment_type="Refund")),
            )
            net_collected = (totals["collected"] or 0) - (totals["refunded"] or 0)
            if payment_type == "Refund" and amount > net_collected:
                raise serializers.ValidationError({"amount": f"Refund cannot exceed the collected balance of {net_collected}."})
            if payment_type != "Refund" and booking.quoted_amount > 0 and net_collected + amount > booking.quoted_amount:
                remaining = max(booking.quoted_amount - net_collected, 0)
                raise serializers.ValidationError({"amount": f"Payment cannot exceed the outstanding balance of {remaining}."})
        if booking and payment_status != "Paid" and payment_type != "Refund":
            pending_duplicate = Payment.objects.filter(
                organization=organization,
                booking=booking,
                payment_type=payment_type,
            ).exclude(status="Paid")
            if self.instance:
                pending_duplicate = pending_duplicate.exclude(pk=self.instance.pk)
            if pending_duplicate.exists():
                raise serializers.ValidationError({"payment_type": "A pending due date already exists for this payment milestone. Edit the existing entry instead."})
        duplicate_fields = ("booking", "customer", "amount", "payment_type", "status", "payment_mode", "received_by", "notes", "due_date")
        defaults = {"booking": None, "customer": None, "amount": None, "payment_type": "Advance", "status": "Pending", "payment_mode": "", "received_by": "", "notes": "", "due_date": None}
        identity = {field: attrs.get(field, getattr(self.instance, field, defaults[field])) for field in duplicate_fields}
        duplicate = Payment.objects.filter(organization=organization, **identity)
        if self.instance:
            duplicate = duplicate.exclude(pk=self.instance.pk)
        if duplicate.exists():
            raise serializers.ValidationError({"detail": "This payment entry already exists. Change the payment details before saving."})
        return attrs


class PaymentViewSet(OrganizationScopedViewSet):
    queryset = Payment.objects.select_related("booking", "customer").all().order_by("-created_at")
    serializer_class = PaymentSerializer
    permission_classes = (AccountsAccessPermission,)
    throttle_scope = "payment_write"
    filterset_fields = {"status": ["exact", "in"], "payment_type": ["exact", "in"], "payment_mode": ["exact", "icontains"], "due_date": ["exact", "gte", "lte"], "paid_at": ["gte", "lte"], "amount": ["gte", "lte"], "booking": ["exact"], "customer": ["exact"], "created_at": ["gte", "lte"]}
    search_fields = ("booking__booking_code", "customer__name", "customer__phone", "received_by", "notes")
    ordering_fields = ("created_at", "amount", "due_date", "paid_at")

    def perform_create(self, serializer):
        paid_at = serializer.validated_data.get("paid_at")
        payment = None
        if serializer.validated_data.get("status") == "Paid" and not paid_at:
            payment = serializer.save(organization=self.request.user.organization, paid_at=timezone.now())
        else:
            payment = super().perform_create(serializer)
        try:
            from apps.notifications.models import broadcast
            broadcast(
                organization=self.request.user.organization,
                title=f"Payment recorded: ₹{payment.amount}",
                body=f"{payment.payment_type} for {payment.customer.name if payment.customer else 'customer'} ({payment.booking.booking_code if payment.booking else '—'}).",
                level="success",
                category="payments",
                link="/accounts",
                payload={"payment_id": payment.id},
            )
        except Exception:
            pass
        return payment

    def perform_update(self, serializer):
        status_value = serializer.validated_data.get("status", serializer.instance.status)
        paid_at = serializer.validated_data.get("paid_at", serializer.instance.paid_at)
        if status_value == "Paid" and not paid_at:
            serializer.save(paid_at=timezone.now())
        else:
            serializer.save()

    def get_throttles(self):
        if self.action in {"import"}:
            self.throttle_scope = "import"
        elif self.action in {"export"}:
            self.throttle_scope = "export"
        return super().get_throttles()

    @action(detail=False, methods=["get"])
    def export(self, request):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Payments"
        fields = ["booking_code", "client_name", "amount", "payment_type", "status", "payment_mode", "received_by", "due_date", "paid_at", "notes"]
        sheet.append(fields)
        for payment in self.get_queryset():
            sheet.append([payment.booking.booking_code, payment.customer.name, payment.amount, payment.payment_type, payment.status, payment.payment_mode, payment.received_by, payment.due_date, payment.paid_at.replace(tzinfo=None) if payment.paid_at else None, payment.notes])
        stream = BytesIO()
        workbook.save(stream)
        response = HttpResponse(stream.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="payments.xlsx"'
        return response

    @action(detail=False, methods=["post"], url_path="import")
    @transaction.atomic
    def import_payments(self, request):
        upload = request.FILES.get("file")
        if not upload:
            return Response({"detail": "Choose an Excel file."}, status=400)
        worksheet_rows = load_workbook(upload, data_only=True).active.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(worksheet_rows)]
        created, skipped = 0, 0
        for values in worksheet_rows:
            row = dict(zip(headers, values))
            booking = Booking.objects.filter(organization=request.user.organization, booking_code=row.get("booking_code")).first()
            if not booking or not row.get("amount"):
                skipped += 1
                continue
            payload = {key: row.get(key) for key in ("amount", "payment_type", "status", "payment_mode", "received_by", "due_date", "paid_at", "notes") if row.get(key) not in (None, "")}
            if payload.get("due_date") and hasattr(payload["due_date"], "strftime"):
                payload["due_date"] = payload["due_date"].strftime("%Y-%m-%d")
            for field in ("paid_at",):
                if payload.get(field) and hasattr(payload[field], "isoformat"):
                    payload[field] = payload[field].isoformat()
            payload.update(booking=booking.id, customer=booking.customer_id)
            serializer = self.get_serializer(data=payload)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            created += 1
        return Response({"created": created, "skipped": skipped}, status=201)


class PaymentReminderSerializer(serializers.ModelSerializer):
    class Meta:
        model = PaymentReminder
        fields = "__all__"
        read_only_fields = ("organization", "reminder_date")
        validators = []

    def validate(self, attrs):
        organization = self.context["request"].user.organization
        booking = attrs.get("booking", getattr(self.instance, "booking", None))
        customer = attrs.get("customer", getattr(self.instance, "customer", None))
        if not organization:
            raise serializers.ValidationError("An organization is required.")
        if booking and booking.organization_id != organization.id:
            raise serializers.ValidationError({"booking": "Invalid booking."})
        if customer and customer.organization_id != organization.id:
            raise serializers.ValidationError({"customer": "Invalid customer."})
        if booking and customer and booking.customer_id != customer.id:
            raise serializers.ValidationError({"customer": "Customer does not match the booking."})
        return attrs


class PaymentReminderViewSet(OrganizationScopedViewSet):
    queryset = PaymentReminder.objects.select_related("booking", "customer").all().order_by("-reminder_date")
    serializer_class = PaymentReminderSerializer
    permission_classes = (AccountsAccessPermission,)
    throttle_scope = "payment_write"
    filterset_fields = {"payment_type": ["exact", "in"], "action": ["exact", "in"], "reminder_date": ["exact", "gte", "lte"], "next_followup_date": ["exact", "gte", "lte"], "booking": ["exact"], "customer": ["exact"]}
    search_fields = ("booking__booking_code", "customer__name", "notes")
    ordering_fields = ("reminder_date", "next_followup_date", "outstanding_amount")

    def create(self, request, *args, **kwargs):
        existing = PaymentReminder.objects.filter(
            organization=request.user.organization,
            booking_id=request.data.get("booking"),
            payment_type=request.data.get("payment_type"),
            reminder_date=timezone.localdate(),
        ).first()
        if existing:
            serializer = self.get_serializer(existing, data=request.data, partial=True)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response(serializer.data)
        return super().create(request, *args, **kwargs)
