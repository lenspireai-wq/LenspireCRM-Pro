from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from rest_framework import serializers
from rest_framework.decorators import action
from rest_framework.response import Response
from apps.core.api import OrganizationScopedViewSet
from apps.core.permissions import OperationsAccessPermission
from .models import CalendarEvent, PhotographerDetail


CREW_FIELDS = ("photo", "video", "candid", "cinematic", "drone", "assistant", "bts")
PENDING_CREW_VALUES = {"", "X", "XX"}


def automatic_event_status(values, current_status: str | None = None) -> str:
    """Derive the operational status from date and assignment readiness."""
    current_status = current_status or str(getattr(values, "status", "Scheduled"))
    if current_status == "Cancelled":
        return "Cancelled"

    def value(field, default=""):
        if isinstance(values, dict):
            return values.get(field, default)
        return getattr(values, field, default)

    start_date = value("start_date", None)
    today = timezone.localdate()
    if start_date and start_date < today:
        return "Completed"
    if start_date == today:
        return "In Progress"
    if not start_date:
        return "Scheduled"

    crew_values = [str(value(field) or "").strip() for field in CREW_FIELDS]
    actual_crew = [item for item in crew_values if item.upper() != "NA"]
    crew_resolved = all(
        item.upper() not in PENDING_CREW_VALUES
        and (item.upper() == "NA" or len("".join(filter(str.isdigit, item))) >= 10)
        for item in crew_values
    )
    details_ready = all(
        (
            str(value("city") or "").strip(),
            value("start_time", None),
            str(value("notes") or "").strip(),
            actual_crew,
        )
    )
    return "Confirmed" if crew_resolved and details_ready else "Scheduled"


class CalendarEventSerializer(serializers.ModelSerializer):
    class Meta:
        model = CalendarEvent
        fields = "__all__"
        read_only_fields = ("organization",)

    def validate(self, attrs):
        request = self.context["request"]
        date_status = attrs.get("date_status", getattr(self.instance, "date_status", "Confirmed"))
        start_date = attrs.get("start_date", getattr(self.instance, "start_date", None))
        tbd_month = attrs.get("tbd_month", getattr(self.instance, "tbd_month", ""))
        if date_status == "Confirmed" and not start_date:
            raise serializers.ValidationError({"start_date": "A confirmed event needs a date."})
        if date_status == "TBD Month" and not tbd_month:
            raise serializers.ValidationError({"tbd_month": "Select the expected month."})
        current_status = attrs.get("status", getattr(self.instance, "status", "Scheduled"))
        lifecycle_values = {
            field: attrs.get(field, getattr(self.instance, field, None))
            for field in ("start_date", "start_time", "city", "notes", *CREW_FIELDS)
        }
        attrs["status"] = automatic_event_status(lifecycle_values, current_status)
        identity_fields = ("client_name", "contact_no", "event_type", "start_date", "start_time", "tbd_month")
        defaults = {"client_name": "", "contact_no": "", "event_type": "Shoot", "start_date": None, "start_time": None, "tbd_month": ""}
        identity = {field: attrs.get(field, getattr(self.instance, field, defaults[field])) for field in identity_fields}
        duplicate = CalendarEvent.objects.filter(organization=request.user.organization, **identity)
        if self.instance:
            duplicate = duplicate.exclude(pk=self.instance.pk)
        if duplicate.exists():
            raise serializers.ValidationError({"detail": "This event already exists. Change the client, contact, event type, date, or time before saving."})
        return attrs

class CalendarEventViewSet(OrganizationScopedViewSet):
    queryset = CalendarEvent.objects.all().order_by("start_date", "start_time")
    serializer_class = CalendarEventSerializer
    permission_classes = (OperationsAccessPermission,)
    filterset_fields = {"status": ["exact", "in"], "event_type": ["exact", "in"], "date_status": ["exact", "in"], "city": ["exact", "icontains"], "start_date": ["exact", "gte", "lte"], "assigned_user": ["exact"], "tbd_month": ["exact", "icontains"]}
    search_fields = ("title", "client_name", "couple_name", "contact_no", "city", "notes", "handled_by", "photo", "video", "candid", "cinematic", "drone", "assistant", "bts")
    ordering_fields = ("start_date", "start_time", "status")

    def get_queryset(self):
        queryset = super().get_queryset()
        changed = []
        for event in queryset.exclude(status="Cancelled"):
            next_status = automatic_event_status(event)
            if event.status != next_status:
                event.status = next_status
                changed.append(event)
        if changed:
            CalendarEvent.objects.bulk_update(changed, ("status",))
        return queryset

    @action(detail=False, methods=["get"])
    def export(self, request):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Upcoming Events"
        fields = ["title", "client_name", "event_type", "start_date", "start_time", "city", "status", "handled_by", "couple_name", "contact_no", "photo", "video", "candid", "cinematic", "drone", "assistant", "bts", "notes"]
        sheet.append(fields)
        for event in self.get_queryset():
            sheet.append([getattr(event, field) for field in fields])
        stream = BytesIO()
        workbook.save(stream)
        response = HttpResponse(stream.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="upcoming-events.xlsx"'
        return response

    @action(detail=False, methods=["post"], url_path="import")
    def import_events(self, request):
        upload = request.FILES.get("file")
        if not upload:
            return Response({"detail": "Choose an Excel file."}, status=400)
        sheet = load_workbook(upload, data_only=True).active
        worksheet_rows = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(worksheet_rows)]
        created = 0
        for values in worksheet_rows:
            row = dict(zip(headers, values))
            if not row.get("title"):
                continue
            editable_fields = {field.name for field in CalendarEvent._meta.fields if field.editable and not field.auto_created} - {"id", "organization", "created_at", "updated_at"}
            payload = {key: row.get(key) for key in headers if key in editable_fields}
            if payload.get("start_date") and hasattr(payload["start_date"], "strftime"):
                payload["start_date"] = payload["start_date"].strftime("%Y-%m-%d")
            for field in ("start_time", "end_time"):
                if payload.get(field) and hasattr(payload[field], "strftime"):
                    payload[field] = payload[field].strftime("%H:%M:%S")
            serializer = self.get_serializer(data=payload)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            created += 1
        return Response({"created": created}, status=201)


class PhotographerDetailSerializer(serializers.ModelSerializer):
    class Meta:
        model = PhotographerDetail
        fields = "__all__"
        read_only_fields = ("organization",)

    def validate(self, attrs):
        request = self.context["request"]
        name = str(attrs.get("name", getattr(self.instance, "name", ""))).strip()
        mobile = "".join(character for character in str(attrs.get("mobile", getattr(self.instance, "mobile", ""))) if character.isdigit())[-10:]
        candidates = PhotographerDetail.objects.filter(organization=request.user.organization)
        if self.instance:
            candidates = candidates.exclude(pk=self.instance.pk)
        duplicate = next(
            (
                person
                for person in candidates.only("id", "name", "mobile")
                if (mobile and "".join(character for character in person.mobile if character.isdigit())[-10:] == mobile)
                or (not mobile and person.name.strip().casefold() == name.casefold())
            ),
            None,
        )
        if duplicate:
            field = "mobile" if mobile else "name"
            raise serializers.ValidationError({field: f"This photographer already exists as {duplicate.name}."})
        return attrs


class PhotographerDetailViewSet(OrganizationScopedViewSet):
    queryset = PhotographerDetail.objects.all()
    serializer_class = PhotographerDetailSerializer
    permission_classes = (OperationsAccessPermission,)
    filterset_fields = {"status": ["exact", "in"], "work": ["exact", "icontains"], "living_in": ["exact", "icontains"]}
    search_fields = ("name", "mobile", "work", "living_in")
    ordering_fields = ("name", "status")
