from io import BytesIO

from django.test import TestCase
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.core.models import Organization
from apps.accounts.models import Payment
from apps.operations.models import CalendarEvent
from apps.sales.models import Booking, Customer, Lead
from apps.users.models import User
from .models import ProductionActivity, ProductionDeliverable, ProductionJob


class ProductionApiTests(TestCase):
    def setUp(self):
        self.organization = Organization.objects.create(name="Studio", slug="studio")
        self.user = User.objects.create_user(username="producer", organization=self.organization)
        self.lead = Lead.objects.create(
            organization=self.organization,
            lead_code="L1",
            name="Client",
            event_type="Wedding",
            status="Confirmed",
            total_closing="1000.00",
        )
        self.customer = Customer.objects.create(
            organization=self.organization,
            customer_code="C1",
            name="Client",
            lead=self.lead,
        )
        self.booking = Booking.objects.create(
            organization=self.organization,
            customer=self.customer,
            lead=self.lead,
            booking_code="B1",
            event_type="Wedding",
            event_date="2026-10-10",
            quoted_amount="1000.00",
        )
        CalendarEvent.objects.create(
            organization=self.organization,
            booking=self.booking,
            customer=self.customer,
            title="Wedding",
            event_type="Wedding",
            start_date="2026-10-10",
            status="Completed",
        )
        Payment.objects.create(
            organization=self.organization,
            booking=self.booking,
            customer=self.customer,
            amount="900.00",
            payment_type="Wedding Day",
            status="Paid",
        )
        self.job = ProductionJob.objects.create(
            organization=self.organization,
            customer=self.customer,
            booking=self.booking,
        )
        self.client = APIClient()
        self.client.force_authenticate(self.user)

    def listed_jobs(self):
        response = self.client.get("/api/production/")
        self.assertEqual(response.status_code, 200, response.data)
        return response.data.get("results", response.data)

    def test_secure_client_portal_link_feedback_and_revocation(self):
        deliverable = ProductionDeliverable.objects.create(
            organization=self.organization,
            job=self.job,
            name="Wedding Gallery",
            status="Sent to Client",
            drive_link="https://drive.google.com/example",
        )
        generated = self.client.post(
            "/api/client-portal/access/",
            {"booking": self.booking.id, "expiry_days": 60},
            format="json",
        )
        self.assertEqual(generated.status_code, 200, generated.data)
        token = generated.data["url"].rstrip("/").split("/")[-1]
        public = APIClient()
        portal = public.get(f"/api/client-portal/{token}/")
        self.assertEqual(portal.status_code, 200, portal.data)
        self.assertEqual(portal.data["booking"]["code"], "B1")
        self.assertEqual(len(portal.data["deliverables"]), 1)
        feedback = public.post(
            f"/api/client-portal/{token}/",
            {"deliverable": deliverable.id, "action": "approve"},
            format="json",
        )
        self.assertEqual(feedback.status_code, 200, feedback.data)
        deliverable.refresh_from_db()
        self.assertEqual(deliverable.status, "Client Approved")
        self.job.refresh_from_db()
        self.assertEqual(self.job.client_approval_status, "Approved")
        self.assertTrue(self.job.activities.filter(activity_type="Client Portal Feedback").exists())
        blocked_delivery = self.client.patch(
            f"/api/production/{self.job.id}/", {"delivery_status": "Delivered"}, format="json"
        )
        self.assertEqual(blocked_delivery.status_code, 400)
        Payment.objects.create(organization=self.organization, booking=self.booking, customer=self.customer, amount="100.00", payment_type="Final Delivery", status="Paid")
        delivered = self.client.patch(
            f"/api/production/{self.job.id}/",
            {"delivery_status": "Delivered", "delivery_method": "Google Drive", "photo_delivery_status": "Delivered", "video_delivery_status": "Not Applicable", "album_delivery_status": "Not Applicable"},
            format="json",
        )
        self.assertEqual(delivered.status_code, 200, delivered.data)
        self.assertEqual(public.get(f"/api/client-portal/{token}/").status_code, 401)
        revoked = self.client.delete(
            "/api/client-portal/access/", {"booking": self.booking.id}, format="json"
        )
        self.assertEqual(revoked.status_code, 200)

    def test_client_invitation_password_login_reset_disable_and_audit(self):
        invited = self.client.post(
            "/api/client-portal/invitations/",
            {"booking": self.booking.id, "name": "Test Client", "email": "client@example.com", "mobile": "919999999999"},
            format="json",
        )
        self.assertEqual(invited.status_code, 200, invited.data)
        self.assertIn("wa.me/919999999999", invited.data["whatsapp_url"])
        whatsapp = self.client.post(
            "/api/client-portal/whatsapp/",
            {"booking": self.booking.id, "message_type": "payment_reminder", "event": "prepared"},
            format="json",
        )
        self.assertEqual(whatsapp.status_code, 200, whatsapp.data)
        self.assertIn("Balance: ₹100.00", whatsapp.data["message"])
        self.assertIn("https://wa.me/?text=", whatsapp.data["whatsapp_url"])
        copied = self.client.put("/api/client-portal/invitations/", {"booking": self.booking.id}, format="json")
        self.assertEqual(copied.status_code, 200)
        invite_token = invited.data["url"].rstrip("/").split("/")[-1]
        public = APIClient()
        setup = public.post("/api/client-portal/auth/setup/", {"token": invite_token, "password": "Client-pass-123"}, format="json")
        self.assertEqual(setup.status_code, 200, setup.data)
        self.assertEqual(public.post("/api/client-portal/auth/setup/", {"token": invite_token, "password": "Client-pass-123"}, format="json").status_code, 401)
        self.assertEqual(public.get(f"/api{setup.data['portal_url']}/").status_code, 200)
        login = public.post("/api/client-portal/auth/login/", {"studio": "studio", "email": "client@example.com", "password": "Client-pass-123"}, format="json")
        self.assertEqual(login.status_code, 200, login.data)
        status_view = self.client.get(f"/api/client-portal/access/?booking={self.booking.id}")
        user = status_view.data["portal_users"][0]
        actions = {item["action"] for item in status_view.data["activities"]}
        self.assertTrue({"Client Invited", "Password Created", "Client Login", "WhatsApp Message Prepared", "WhatsApp Invitation Copied"}.issubset(actions))
        disabled = self.client.patch("/api/client-portal/invitations/", {"user": user["id"], "action": "disable"}, format="json")
        self.assertEqual(disabled.status_code, 200)
        self.assertEqual(public.get(f"/api{login.data['portal_url']}/").status_code, 401)
        reset = self.client.patch("/api/client-portal/invitations/", {"user": user["id"], "action": "reset"}, format="json")
        self.assertEqual(reset.status_code, 200)
        self.assertIn("/client-portal/setup/", reset.data["url"])
    def test_edit_queue_requires_completed_event_and_ninety_percent_payment(self):
        self.assertEqual(len(self.listed_jobs()), 1)
        payment = Payment.objects.get(booking=self.booking)
        payment.amount = "899.00"
        payment.save(update_fields=["amount"])
        self.assertEqual(len(self.listed_jobs()), 0)
        payment.amount = "900.00"
        payment.save(update_fields=["amount"])
        event = CalendarEvent.objects.get(booking=self.booking)
        event.status = "Scheduled"
        event.save(update_fields=["status"])
        self.assertEqual(len(self.listed_jobs()), 0)

    def test_combined_package_requires_any_qualifying_completion_and_fifty_percent(self):
        CalendarEvent.objects.create(
            organization=self.organization,
            booking=self.booking,
            customer=self.customer,
            title="Engagement",
            event_type="Engagement",
            status="Scheduled",
        )
        CalendarEvent.objects.create(
            organization=self.organization,
            booking=self.booking,
            customer=self.customer,
            title="Pre-Wedding",
            event_type="Pre-Wedding",
            status="Completed",
        )
        CalendarEvent.objects.filter(
            booking=self.booking, event_type="Wedding"
        ).update(status="Scheduled")
        Payment.objects.filter(booking=self.booking).delete()
        Payment.objects.create(
            organization=self.organization,
            booking=self.booking,
            customer=self.customer,
            amount="100.00",
            payment_type="Other",
            status="Paid",
        )
        first_shoot = Payment.objects.create(
            organization=self.organization,
            booking=self.booking,
            customer=self.customer,
            amount="400.00",
            payment_type="Wedding Day",
            status="Paid",
        )
        self.assertEqual(len(self.listed_jobs()), 1)
        first_shoot.amount = "399.00"
        first_shoot.save(update_fields=["amount"])
        self.assertEqual(len(self.listed_jobs()), 0)

    def test_job_includes_connected_booking_details(self):
        response = self.client.get(f"/api/production/{self.job.id}/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["client_name"], "Client")
        self.assertEqual(response.data["booking_code"], "B1")
        self.assertEqual(response.data["event_type"], "Wedding")

    def test_marking_delivered_records_delivery_time(self):
        Payment.objects.create(
            organization=self.organization,
            booking=self.booking,
            customer=self.customer,
            amount="100.00",
            payment_type="Final Delivery",
            status="Paid",
        )
        response = self.client.patch(
            f"/api/production/{self.job.id}/",
            {
                "delivery_status": "Delivered",
                "client_approval_status": "Approved",
                "delivery_method": "Google Drive",
                "photo_delivery_status": "Delivered",
                "video_delivery_status": "Delivered",
                "album_delivery_status": "Not Applicable",
            },
        )
        self.assertEqual(response.status_code, 200, response.data)
        self.assertIsNotNone(response.data["delivered_at"])
        self.assertIsNotNone(response.data["client_approved_at"])
        self.assertEqual(response.data["stage"], "Delivered")

    def test_delivery_requires_approval_checklist_and_method(self):
        response = self.client.patch(
            f"/api/production/{self.job.id}/", {"delivery_status": "Delivered"}
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("client_approval_status", response.data)

    def test_editor_list_and_assignment_are_organization_scoped(self):
        editor = User.objects.create_user(
            username="editor",
            display_name="Studio Editor",
            mobile="9876543210",
            organization=self.organization,
            role="Editor",
        )
        response = self.client.get("/api/production/editors/")
        self.assertEqual(response.status_code, 200)
        self.assertIn("Studio Editor", [item["name"] for item in response.data])
        assigned = self.client.patch(
            f"/api/production/{self.job.id}/", {"editor": editor.id}
        )
        self.assertEqual(assigned.status_code, 200, assigned.data)
        self.assertEqual(assigned.data["editor_name"], "Studio Editor")
        self.assertEqual(assigned.data["editor_mobile"], "9876543210")
        self.assertTrue(
            ProductionActivity.objects.filter(
                job=self.job, activity_type="Editor Assignment"
            ).exists()
        )

    def test_status_changes_are_added_to_activity_history(self):
        response = self.client.patch(
            f"/api/production/{self.job.id}/",
            {"stage": "Editing", "editing_status": "In Progress"},
        )
        self.assertEqual(response.status_code, 200, response.data)
        self.assertEqual(
            set(
                ProductionActivity.objects.filter(job=self.job).values_list(
                    "activity_type", flat=True
                )
            ),
            {"Stage Change", "Editing Status"},
        )
        self.assertEqual(len(response.data["activities"]), 2)

    def test_deliverable_assignment_appears_only_for_assigned_editor(self):
        editor = User.objects.create_user(
            username="assigned-editor",
            display_name="Assigned Editor",
            organization=self.organization,
            role="Editor",
            department_access={"production": "full"},
        )
        response = self.client.patch(
            f"/api/production/{self.job.id}/",
            {
                "deliverables": [
                    {
                        "name": "Cinematic Highlight",
                        "enabled": True,
                        "quantity": 1,
                        "editor": editor.id,
                        "due_date": "2026-10-20",
                        "priority": "High",
                        "status": "Assigned",
                    }
                ]
            },
            format="json",
        )
        self.assertEqual(response.status_code, 200, response.data)
        task = response.data["deliverables"][0]
        self.assertEqual(task["editor_name"], "Assigned Editor")

        editor_client = APIClient()
        editor_client.force_authenticate(editor)
        jobs = editor_client.get("/api/production/")
        self.assertEqual(jobs.status_code, 200)
        editor_jobs = jobs.data.get("results", jobs.data)
        self.assertEqual(len(editor_jobs), 1)
        blocked = editor_client.patch(
            f"/api/production/{self.job.id}/", {"stage": "Editing"}
        )
        self.assertEqual(blocked.status_code, 403)
        started = editor_client.patch(
            f"/api/production/{self.job.id}/deliverables/{task['id']}/",
            {"status": "In Progress"},
            format="json",
        )
        self.assertEqual(started.status_code, 200, started.data)
        submitted = editor_client.patch(
            f"/api/production/{self.job.id}/deliverables/{task['id']}/",
            {
                "status": "Submitted for Review",
                "drive_link": "https://drive.google.com/example",
            },
            format="json",
        )
        self.assertEqual(submitted.status_code, 200, submitted.data)
        self.assertEqual(submitted.data["status"], "Submitted for Review")
        approved = self.client.patch(
            f"/api/production/{self.job.id}/",
            {
                "deliverables": [
                    {
                        "name": "Cinematic Highlight",
                        "enabled": True,
                        "quantity": 1,
                        "editor": editor.id,
                        "priority": "High",
                        "status": "Approved",
                        "drive_link": "https://drive.google.com/example",
                    }
                ]
            },
            format="json",
        )
        self.assertEqual(approved.status_code, 200, approved.data)
        self.assertEqual(approved.data["deliverables"][0]["status"], "Approved")
        self.assertTrue(
            ProductionActivity.objects.filter(
                job=self.job,
                activity_type="Deliverable Status",
                description__contains="Approved",
            ).exists()
        )

    def test_revision_decision_requires_instructions(self):
        editor = User.objects.create_user(
            username="revision-editor",
            organization=self.organization,
            role="Editor",
        )
        response = self.client.patch(
            f"/api/production/{self.job.id}/",
            {
                "deliverables": [
                    {
                        "name": "Reels",
                        "enabled": True,
                        "editor": editor.id,
                        "status": "Revision Required",
                        "revision_notes": "",
                    }
                ]
            },
            format="json",
        )
        self.assertEqual(response.status_code, 400)

    def test_only_one_overdue_reminder_is_logged_each_day(self):
        self.job.due_date = "2020-01-01"
        self.job.save(update_fields=["due_date"])
        first = self.client.post(f"/api/production/{self.job.id}/reminder/")
        second = self.client.post(f"/api/production/{self.job.id}/reminder/")
        self.assertEqual(first.status_code, 201, first.data)
        self.assertEqual(second.status_code, 200, second.data)
        self.assertEqual(
            ProductionActivity.objects.filter(
                job=self.job, activity_type="Overdue Reminder"
            ).count(),
            1,
        )

    def test_reminder_rejects_a_job_that_is_not_overdue(self):
        response = self.client.post(f"/api/production/{self.job.id}/reminder/")
        self.assertEqual(response.status_code, 400)

    def test_filtered_production_export(self):
        self.job.stage = "Editing"
        self.job.save(update_fields=["stage"])
        response = self.client.get("/api/production/export/?stage=Editing")
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))
        self.assertEqual(rows[0][0:3], ("Client", "Booking", "Event"))
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[1][0:3], ("Client", "B1", "Wedding"))

        empty = self.client.get("/api/production/export/?stage=Delivered")
        empty_rows = list(
            load_workbook(BytesIO(empty.content), data_only=True)
            .active.iter_rows(values_only=True)
        )
        self.assertEqual(len(empty_rows), 1)

    def test_activity_history_and_filtered_export(self):
        ProductionActivity.objects.create(
            organization=self.organization,
            job=self.job,
            activity_type="Stage Change",
            description="Stage changed from Shoot Planning to Editing.",
            performed_by="Producer",
        )
        history = self.client.get(
            "/api/production/activity-history/?activity_type=Stage%20Change"
        )
        self.assertEqual(history.status_code, 200)
        self.assertEqual(len(history.data), 1)
        self.assertEqual(history.data[0]["client_name"], "Client")
        self.assertEqual(history.data[0]["booking_code"], "B1")

        response = self.client.get(
            "/api/production/activity-export/?activity_type=Stage%20Change"
        )
        self.assertEqual(response.status_code, 200)
        rows = list(
            load_workbook(BytesIO(response.content), data_only=True)
            .active.iter_rows(values_only=True)
        )
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[1][1:5], ("Client", "B1", "Wedding", "Stage Change"))

    def test_rejects_booking_from_another_organization(self):
        other = Organization.objects.create(name="Other", slug="other")
        customer = Customer.objects.create(
            organization=other, customer_code="C2", name="Other Client"
        )
        booking = Booking.objects.create(
            organization=other,
            customer=customer,
            booking_code="B2",
            event_type="Wedding",
        )
        response = self.client.patch(
            f"/api/production/{self.job.id}/",
            {"booking": booking.id, "customer": customer.id},
        )
        self.assertEqual(response.status_code, 400)
        other_editor = User.objects.create_user(username="other-editor", organization=other)
        response = self.client.patch(
            f"/api/production/{self.job.id}/", {"editor": other_editor.id}
        )
        self.assertEqual(response.status_code, 400)
